"""
拿 openpyxl（别人写的解析器）读我们手写的 xlsx。
**自己写自己读是循环论证** —— 文件格式最容易在那儿栽：我以为写对了，Excel 说"文件已损坏"。
判据不是"读不读得开"，是**每张 sheet 的名字、行数、以及数字有没有真的是数字**。
数字这条最要紧：写成文本的话，人在 Excel 里选中一列，底下不显示合计，而这个错不报任何东西。
"""
import subprocess
import sys
import tempfile
import os

try:
    import openpyxl
except ImportError:
    print("⏭  没装 openpyxl，这一档没验 —— 装：pip3 install openpyxl")
    print("   （这不等于通过：手写 xlsx 有没有写对，只有真解析器说了算）")
    sys.exit(0)

根 = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
造 = r'''
import { 造xlsx } from 'ROOT/lib/合表.mjs';
import { writeFileSync } from 'node:fs';
writeFileSync(process.argv[2], 造xlsx([
  { 名: '光模块-闵行', 行们: [['型号', '品牌', '在库'], ['QSFP112-400G-DR4', '海光芯创', '47823'], ['OSFP112-800G', '海光芯创', '7480']] },
  { 名: '网卡-闵行', 行们: [['型号', '品牌', '在库'], ['CX6', 'NVIDIA', '4235']] },
  { 名: 'a/b:c*d?e[f]', 行们: [['带非法字符的名字'], ['x']] },
]));
'''

挂 = 0


def 判(条件, 说):
    global 挂
    if not 条件:
        挂 += 1
        print(f"✗ {说}")
    else:
        print(f"✓ {说}")


with tempfile.TemporaryDirectory() as d:
    js = os.path.join(d, "造.mjs")
    xlsx = os.path.join(d, "试.xlsx")
    with open(js, "w", encoding="utf-8") as f:
        # ESM 的相对 import 是相对**模块文件自己**的位置，不是相对 cwd —— 临时目录里写死绝对路径
        f.write(造.replace("ROOT", 根))
    r = subprocess.run(["node", js, xlsx], cwd=根, capture_output=True, text=True)
    if r.returncode != 0:
        print("✗ 造 xlsx 就失败了：" + (r.stderr or "")[:400])
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(xlsx)
    except Exception as e:  # noqa: BLE001
        print(f"✗ **openpyxl 打不开这个文件**：{type(e).__name__}: {e}")
        print("   这就是自己读自己验不出来的那类错 —— 我们这边一点征兆都没有，Excel 那边直接报损坏。")
        sys.exit(1)

    判(len(wb.sheetnames) == 3, f"openpyxl 读到 3 张 sheet（实际 {len(wb.sheetnames)}：{wb.sheetnames}）")
    判(wb.sheetnames[0] == "光模块-闵行", "第一张 sheet 名字对（中文没乱码）")
    判("/" not in wb.sheetnames[2] and ":" not in wb.sheetnames[2],
       f"非法字符被洗掉了（第三张叫「{wb.sheetnames[2]}」）")

    ws = wb["光模块-闵行"]
    判(ws.max_row == 3, f"光模块那张 3 行（表头+2 行数据，实际 {ws.max_row}）")
    判(ws["A1"].value == "型号", f"表头读得对（A1={ws['A1'].value!r}）")
    判(ws["A2"].value == "QSFP112-400G-DR4", f"中文/型号串原样（A2={ws['A2'].value!r}）")

    # **最要紧的一条**：在库那列必须是数字，不是文本
    c = ws["C2"]
    判(isinstance(c.value, (int, float)), f"**在库是数字不是文本**（C2={c.value!r} 类型 {type(c.value).__name__}）")
    判(c.value == 47823, f"数值对（C2={c.value}）")
    判(isinstance(ws["A2"].value, str), "型号还是文本（别把什么都转成数字）")

    ws2 = wb["网卡-闵行"]
    判(ws2["C2"].value == 4235, f"第二张 sheet 的数也对（C2={ws2['C2'].value}）")

print(f"\n检查了 {10} 条判据，失败 {挂}")
sys.exit(1 if 挂 else 0)
